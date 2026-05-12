#!/usr/bin/env python3
"""
Скрипт синхронизации данных агентов с Яндекс.Диска.

Использование:
  python3 scripts/sync_agents.py

Что делает:
  1. Скачивает Excel-файл с Яндекс.Диска по публичной ссылке
  2. Парсит лист «БД» (терминалы и их статусы)
  3. Парсит лист «СВОДНАЯ» (флаги сертификатов и продаж)
  4. Записывает результат в public/agents.json
"""

import json
import sys
import urllib.request
import urllib.parse
from pathlib import Path

# ─── Настройки ────────────────────────────────────────────────────────────────
YANDEX_DISK_PUBLIC_URL = "https://disk.yandex.ru/i/Se7v9O6ME4LyhA"
OUTPUT_PATH = Path(__file__).parent.parent / "public" / "agents.json"

# Терминалы с этими статусами не попадают в панель напоминаний
SKIP_STATUSES = {"добавлен", "заблокирован"}

# ─── Получаем прямую ссылку на скачивание через API Яндекс.Диска ─────────────
def get_download_url(public_url: str) -> str:
    api_url = (
        "https://cloud-api.yandex.net/v1/disk/public/resources/download"
        f"?public_key={urllib.parse.quote(public_url)}"
    )
    req = urllib.request.Request(api_url, headers={"Accept": "application/json"})
    with urllib.request.urlopen(req) as resp:
        data = json.loads(resp.read())
        return data["href"]

# ─── Скачиваем файл во временную папку ───────────────────────────────────────
def download_file(url: str, dest: Path) -> None:
    urllib.request.urlretrieve(url, dest)

# ─── Парсим Excel ─────────────────────────────────────────────────────────────
def parse_excel(path: Path) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        print("Установите openpyxl: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(path)

    # ── Шаг 1: Лист СВОДНАЯ → флаги по агенту ────────────────────────────────
    # Колонки: №, ЮЛ, МТС ID, Terminal ID, Агент ID, GUID,
    #          Ответственный, Комментарии,
    #          Добавлен сертификат, Добавлен сертификат (МТС), Билеты продаются
    summary: dict[int, dict] = {}  # gateId -> flags
    ws_sum = wb["СВОДНАЯ"]
    for row in ws_sum.iter_rows(min_row=2, values_only=True):
        if len(row) < 11:
            continue
        _, name, _, _, agent_id, _, _, comment, cert, cert_mts, tickets = row[:11]
        if not agent_id:
            continue
        try:
            gate_id = int(str(agent_id).strip())
        except ValueError:
            continue
        summary[gate_id] = {
            "certAdded":    bool(cert),
            "certMtsAdded": bool(cert_mts),
            "ticketsSelling": bool(tickets),
            "comment":      str(comment).strip() if comment else None,
        }

    # ── Шаг 2: Лист БД → терминалы + проблемные статусы ─────────────────────
    # Колонки: №, ЮЛ, МТС ID, Регион, Город, Улица, Дом,
    #          Terminal ID, Агент ID, GUID, Ответственный, Комментарии, Статус
    agents: dict[int, dict] = {}
    ws_bd = wb["БД"]
    for row in ws_bd.iter_rows(min_row=2, values_only=True):
        if len(row) < 13:
            continue
        _, name, _, _, _, _, _, terminal_id, agent_id, guid, _, _, status = row[:13]
        if not name or not agent_id:
            continue
        gate_id = int(agent_id)

        if gate_id not in agents:
            flags = summary.get(gate_id, {
                "certAdded": False,
                "certMtsAdded": False,
                "ticketsSelling": False,
                "comment": None,
            })
            agents[gate_id] = {
                "name":              str(name).strip(),
                "gateId":            gate_id,
                "guid1c":            str(guid).strip() if guid else "",
                "salesChannel1cbu":  "",
                "salesChannelReports": "",
                "terminalIds":       [],
                "certAdded":         flags["certAdded"],
                "certMtsAdded":      flags["certMtsAdded"],
                "ticketsSelling":    flags["ticketsSelling"],
                "comment":           flags["comment"],
                "problemTerminals":  [],  # [{id, status}] — только не OK
            }

        if terminal_id:
            tid = int(terminal_id)
            if tid not in agents[gate_id]["terminalIds"]:
                agents[gate_id]["terminalIds"].append(tid)

            # Добавляем в проблемные если статус не в списке пропускаемых
            status_str = str(status).strip() if status else ""
            if status_str and status_str not in SKIP_STATUSES:
                # Не дублируем
                existing_ids = [p["id"] for p in agents[gate_id]["problemTerminals"]]
                if tid not in existing_ids:
                    agents[gate_id]["problemTerminals"].append({
                        "id": tid,
                        "status": status_str,
                    })

    # ── Финальная сортировка ──────────────────────────────────────────────────
    result = sorted(agents.values(), key=lambda x: x["gateId"])
    for a in result:
        a["terminalIds"].sort()
        a["problemTerminals"].sort(key=lambda x: x["id"])

    return result

# ─── Точка входа ─────────────────────────────────────────────────────────────
def main():
    import tempfile

    print("📥 Получаем ссылку на скачивание с Яндекс.Диска…")
    download_url = get_download_url(YANDEX_DISK_PUBLIC_URL)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)

    print("📦 Скачиваем файл…")
    download_file(download_url, tmp_path)

    print("📊 Парсим Excel…")
    data = parse_excel(tmp_path)
    tmp_path.unlink()

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

    problems = sum(
        1 for a in data
        if not a["certAdded"] or not a["certMtsAdded"]
        or not a["ticketsSelling"] or a["problemTerminals"]
    )
    print(f"✅ Готово: {len(data)} агентов, {problems} требуют внимания → {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
